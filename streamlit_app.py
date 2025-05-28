import streamlit as st
import pandas as pd
from io import BytesIO
import re
from collections import defaultdict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import io

@st.cache_data
def convert_df(df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
    output.seek(0)
    return output

st.set_page_config(page_title="Sales Reports Dashboard", layout="wide")

# Define tab structure
tab1, tab2, tab3 = st.tabs(["📋 Report 1: Sales Summary with RBM/BDM", "📋 Report 2: Store Summary","📊 OSG DATA Mapping"])

# --------------------------- REPORT 1 TAB ---------------------------
with tab1:
    st.header("Report 1: Sales Summary with RBM and BDM")
    st.markdown("""
    Upload the following three files:
    - **Book1.xlsx** (Sales data with DATE, Store, QUANTITY, AMOUNT)
    - **Future store list.xlsx** (Reference list of stores)
    - **RBM and BDM file** (Mapping of Store to RBM and BDM)
    """)

    book1_file = st.file_uploader("Upload Book1.xlsx", type=["xlsx"], key="r1_book1")
    future_store_file = st.file_uploader("Upload future store list.xlsx", type=["xlsx"], key="r1_future")
    rbm_bdm_file = st.file_uploader("Upload RBM and BDM file", type=["xlsx"], key="r1_rbm")

    if book1_file and future_store_file and rbm_bdm_file:
        book1_df = pd.read_excel(book1_file)
        future_store_df = pd.read_excel(future_store_file)
        rbm_bdm_df = pd.read_excel(rbm_bdm_file)

        book1_df.rename(columns={'Branch': 'Store'}, inplace=True)
        rbm_bdm_df.rename(columns={'Branch': 'Store'}, inplace=True)

        book1_df['DATE'] = pd.to_datetime(book1_df['DATE'], dayfirst=True, errors='coerce')
        book1_df.dropna(subset=['DATE'], inplace=True)

        today = pd.to_datetime("27-05-2025", dayfirst=True)
        mtd_df = book1_df[book1_df['DATE'].dt.month == today.month]
        today_df = mtd_df[mtd_df['DATE'].dt.date == today.date()]

        today_agg = today_df.groupby('Store', as_index=False).agg({'QUANTITY': 'sum', 'AMOUNT': 'sum'}).rename(
            columns={'QUANTITY': 'FTD Count', 'AMOUNT': 'FTD Amount'})
        mtd_agg = mtd_df.groupby('Store', as_index=False).agg({'QUANTITY': 'sum', 'AMOUNT': 'sum'}).rename(
            columns={'QUANTITY': 'MTD Count', 'AMOUNT': 'MTD Amount'})

        all_store_names = pd.Series(pd.concat([future_store_df['Store'], book1_df['Store']]).unique(), name='Store')
        report_df = pd.DataFrame(all_store_names)

        report_df = report_df.merge(today_agg, on='Store', how='left').merge(mtd_agg, on='Store', how='left')
        report_df[['FTD Count', 'FTD Amount', 'MTD Count', 'MTD Amount']] = report_df[
            ['FTD Count', 'FTD Amount', 'MTD Count', 'MTD Amount']].fillna(0).astype(int)
        report_df = report_df.merge(rbm_bdm_df[['Store', 'RBM', 'BDM']], on='Store', how='left')
        report_df = report_df.sort_values('MTD Amount', ascending=False)

        # Excel report generation
        def generate_report1_excel(df):
            wb = Workbook()
            wb.remove(wb.active)

            def write_sheet(ws, data):
                headers = ['Store', 'FTD Count', 'FTD Amount', 'MTD Count', 'MTD Amount']
                header_fill = PatternFill("solid", fgColor="4F81BD")
                header_font = Font(bold=True, color="FFFFFF")
                data_fill = PatternFill("solid", fgColor="DCE6F1")
                zero_fill = PatternFill("solid", fgColor="F4CCCC")
                total_fill = PatternFill("solid", fgColor="FFD966")
                border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                bottom=Side(style='thin'))

                for r_idx, row in enumerate(dataframe_to_rows(data[headers], index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=r_idx, column=c_idx, value=value)
                        if r_idx == 1:
                            cell.fill = header_fill
                            cell.font = header_font
                        else:
                            ftd, mtd = row[1], row[3]
                            cell.fill = zero_fill if (ftd == 0 or mtd == 0) else data_fill
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center')

                # Total row
                row_idx = ws.max_row + 1
                ws.cell(row=row_idx, column=1, value="TOTAL").fill = total_fill
                for col_idx in range(2, len(headers) + 1):
                    val = int(data[headers[col_idx - 1]].sum())
                    ws.cell(row=row_idx, column=col_idx, value=val).fill = total_fill

            ws_all = wb.create_sheet("All_Stores")
            write_sheet(ws_all, report_df)

            for rbm in report_df['RBM'].dropna().unique():
                ws = wb.create_sheet(rbm[:30])
                write_sheet(ws, report_df[report_df['RBM'] == rbm])

            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)
            return buf

        excel_buf1 = generate_report1_excel(report_df)
        st.download_button("⬇ Download Excel Report 1", data=excel_buf1, file_name="Report1_Sales_Summary.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        st.info("Please upload all three required files.")

# --------------------------- REPORT 2 TAB ---------------------------
with tab2:
    st.header("Report 2: Store Summary Report")
    st.markdown("""
    Upload the following files:
    - **Book1.xlsx** (Sales data with Store, QUANTITY, AMOUNT)
    - **Future store list.xlsx** (Store master list)
    """)

    book2_file = st.file_uploader("Upload Book1.xlsx", type=["xlsx"], key="r2_book1")
    store_list_file = st.file_uploader("Upload future store list.xlsx", type=["xlsx"], key="r2_store_list")

    if book2_file and store_list_file:
        book2_df = pd.read_excel(book2_file)
        future_df = pd.read_excel(store_list_file)

        book2_df.rename(columns={'Branch': 'Store'}, inplace=True)
        agg = book2_df.groupby('Store', as_index=False).agg({'QUANTITY': 'sum', 'AMOUNT': 'sum'})

        all_stores = pd.DataFrame(pd.concat([future_df['Store'], agg['Store']]).unique(), columns=['Store'])
        merged = all_stores.merge(agg, on='Store', how='left')
        merged['QUANTITY'] = merged['QUANTITY'].fillna(0).astype(int)
        merged['AMOUNT'] = merged['AMOUNT'].fillna(0).astype(int)

        merged = merged.sort_values(by='AMOUNT', ascending=False).reset_index(drop=True)
        total = pd.DataFrame([{
            'Store': 'TOTAL',
            'QUANTITY': merged['QUANTITY'].sum(),
            'AMOUNT': merged['AMOUNT'].sum()
        }])
        final_df = pd.concat([merged, total], ignore_index=True)
        final_df.rename(columns={'Store': 'Branch'}, inplace=True)

        def generate_report2_excel(df):
            wb = Workbook()
            ws = wb.active
            ws.title = "Store Report"

            header_fill = PatternFill("solid", fgColor="4F81BD")
            data_fill = PatternFill("solid", fgColor="DCE6F1")
            red_fill = PatternFill("solid", fgColor="F4CCCC")
            total_fill = PatternFill("solid", fgColor="FFD966")
            border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))
            bold_font = Font(bold=True)
            header_font = Font(bold=True, color="FFFFFF")

            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:
                        cell.fill = header_fill
                        cell.font = header_font
                    elif df.loc[r_idx - 2, 'Branch'] == 'TOTAL':
                        cell.fill = total_fill
                        cell.font = bold_font
                    elif df.loc[r_idx - 2, 'AMOUNT'] <= 0:
                        cell.fill = red_fill
                    else:
                        cell.fill = data_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')

            for column_cells in ws.columns:
                length = max(len(str(cell.value)) for cell in column_cells if cell.value)
                ws.column_dimensions[column_cells[0].column_letter].width = length + 2

            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)
            return buf

        excel_buf2 = generate_report2_excel(final_df)
        st.download_button("⬇ Download Excel Report 2", data=excel_buf2,
                           file_name="Report2_Store_Summary.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        st.info("Please upload both required files.")
with tab3:
    st.header("🔍 OSG & PRODUCT Data Mapping")

    osg_file = st.file_uploader("Upload OSG File", type=["xlsx"], key="osg_mapping")
    product_file = st.file_uploader("Upload PRODUCT File", type=["xlsx"], key="product_mapping")

    if osg_file and product_file:
        osg_df = pd.read_excel(osg_file)
        product_df = pd.read_excel(product_file)

        # SKU Mapping
        sku_category_mapping = {
            "Warranty : Water Cooler/Dispencer/Geyser/RoomCooler/Heater": [
                "COOLER", "DISPENCER", "GEYSER", "ROOM COOLER", "HEATER", "WATER HEATER", "WATER DISPENSER"
            ],
            "Warranty : Fan/Mixr/IrnBox/Kettle/OTG/Grmr/Geysr/Steamr/Inductn": [
                "FAN", "MIXER", "IRON BOX", "KETTLE", "OTG", "GROOMING KIT", "GEYSER", "STEAMER", "INDUCTION",
                "CEILING FAN", "TOWER FAN", "PEDESTAL FAN", "INDUCTION COOKER", "ELECTRIC KETTLE", "WALL FAN", "MIXER GRINDER", "CELLING FAN"
            ],
            "AC : EWP : Warranty : AC": ["AC", "AIR CONDITIONER", "AC INDOOR"],
            "HAEW : Warranty : Air Purifier/WaterPurifier": ["AIR PURIFIER", "WATER PURIFIER"],
            "HAEW : Warranty : Dryer/MW/DishW": ["DRYER", "MICROWAVE OVEN", "DISH WASHER", "MICROWAVE OVEN-CONV"],
            "HAEW : Warranty : Ref/WM": [
                "REFRIGERATOR", "WASHING MACHINE", "WASHING MACHINE-TL", "REFRIGERATOR-DC",
                "WASHING MACHINE-FL", "WASHING MACHINE-SA", "REF", "REFRIGERATOR-CBU", "REFRIGERATOR-FF", "WM"
            ],
            "HAEW : Warranty : TV": ["TV", "TV 28 %", "TV 18 %"],
            "TV : TTC : Warranty and Protection : TV": ["TV", "TV 28 %", "TV 18 %"],
            "TV : Spill and Drop Protection": ["TV", "TV 28 %", "TV 18 %"],
            "HAEW : Warranty :Chop/Blend/Toast/Air Fryer/Food Processr/JMG/Induction": [
                "CHOPPER", "BLENDER", "TOASTER", "AIR FRYER", "FOOD PROCESSOR", "JUICER", "INDUCTION COOKER"
            ],
            "HAEW : Warranty : HOB and Chimney": ["HOB", "CHIMNEY"],
            "HAEW : Warranty : HT/SoundBar/AudioSystems/PortableSpkr": [
                "HOME THEATRE", "AUDIO SYSTEM", "SPEAKER", "SOUND BAR", "PARTY SPEAKER"
            ],
            "HAEW : Warranty : Vacuum Cleaner/Fans/Groom&HairCare/Massager/Iron": [
                "VACUUM CLEANER", "FAN", "MASSAGER", "IRON BOX", "CEILING FAN", "TOWER FAN", "PEDESTAL FAN", "WALL FAN", "ROBO VACCUM CLEANER"
            ],
            "AC AMC": ["AC", "AC INDOOR"]
        }

        product_df['Category'] = product_df['Category'].str.upper().fillna('')
        product_df['Model'] = product_df['Model'].fillna('')
        product_df['Customer Mobile'] = product_df['Customer Mobile'].astype(str)
        product_df['Invoice Number'] = product_df['Invoice Number'].astype(str)
        product_df['Item Rate'] = pd.to_numeric(product_df['Item Rate'], errors='coerce')
        product_df['IMEI'] = product_df['IMEI'].astype(str).fillna('')
        product_df['Brand'] = product_df['Brand'].fillna('')
        osg_df['Customer Mobile'] = osg_df['Customer Mobile'].astype(str)

        def extract_price_slab(text):
            match = re.search(r"Slab\s*:\s*(\d+)K-(\d+)K", str(text))
            if match:
                return int(match.group(1)) * 1000, int(match.group(2)) * 1000
            return None, None

        def get_model(row):
            mobile = row['Customer Mobile']
            retailer_sku = str(row['Retailer SKU'])
            invoice = str(row.get('Invoice Number', ''))
            user_products = product_df[product_df['Customer Mobile'] == mobile]

            if user_products.empty:
                return ''
            unique_models = user_products['Model'].dropna().unique()
            if len(unique_models) == 1:
                return unique_models[0]

            mapped_keywords = []
            for sku_key, keywords in sku_category_mapping.items():
                if sku_key in retailer_sku:
                    mapped_keywords = [kw.lower() for kw in keywords]
                    break

            filtered = user_products[user_products['Category'].str.lower().isin(mapped_keywords)]
            if filtered['Model'].nunique() == 1:
                return filtered['Model'].iloc[0]

            slab_min, slab_max = extract_price_slab(retailer_sku)
            if slab_min and slab_max:
                slab_filtered = filtered[(filtered['Item Rate'] >= slab_min) & (filtered['Item Rate'] <= slab_max)]
                if slab_filtered['Model'].nunique() == 1:
                    return slab_filtered['Model'].iloc[0]
                invoice_filtered = slab_filtered[slab_filtered['Invoice Number'].astype(str) == invoice]
                if invoice_filtered['Model'].nunique() == 1:
                    return invoice_filtered['Model'].iloc[0]

            return ''

        osg_df['Model'] = osg_df.apply(get_model, axis=1)
        category_brand_df = product_df[['Customer Mobile', 'Model', 'Category', 'Brand']].drop_duplicates()
        osg_df = osg_df.merge(category_brand_df, on=['Customer Mobile', 'Model'], how='left')

        invoice_pool = defaultdict(list)
        itemrate_pool = defaultdict(list)
        imei_pool = defaultdict(list)

        for _, row in product_df.iterrows():
            key = (row['Customer Mobile'], row['Model'])
            invoice_pool[key].append(row['Invoice Number'])
            itemrate_pool[key].append(row['Item Rate'])
            imei_pool[key].append(row['IMEI'])

        invoice_usage_counter = defaultdict(int)
        itemrate_usage_counter = defaultdict(int)
        imei_usage_counter = defaultdict(int)

        def assign_from_pool(row, pool, counter_dict):
            key = (row['Customer Mobile'], row['Model'])
            values = pool.get(key, [])
            index = counter_dict[key]
            if index < len(values):
                counter_dict[key] += 1
                return values[index]
            return ''

        osg_df['Product Invoice Number'] = osg_df.apply(lambda row: assign_from_pool(row, invoice_pool, invoice_usage_counter), axis=1)
        osg_df['Item Rate'] = osg_df.apply(lambda row: assign_from_pool(row, itemrate_pool, itemrate_usage_counter), axis=1)
        osg_df['IMEI'] = osg_df.apply(lambda row: assign_from_pool(row, imei_pool, imei_usage_counter), axis=1)
        osg_df['Store Code'] = osg_df['Product Invoice Number'].astype(str).apply(
            lambda x: re.search(r'\b([A-Z]{2,})\b', x).group(1) if re.search(r'\b([A-Z]{2,})\b', x) else ''
        )

        def extract_warranty_duration(sku):
            sku = str(sku)
            match = re.search(r'Dur\s*:\s*(\d+)\+(\d+)', sku)
            if match:
                return int(match.group(1)), int(match.group(2))
            match = re.search(r'(\d+)\+(\d+)\s*SDP-(\d+)', sku)
            if match:
                return int(match.group(1)), f"{match.group(3)}P+{match.group(2)}W"
            match = re.search(r'Dur\s*:\s*(\d+)', sku)
            if match:
                return 1, int(match.group(1))
            match = re.search(r'(\d+)\+(\d+)', sku)
            if match:
                return int(match.group(1)), int(match.group(2))
            return '', ''

        osg_df[['Manufacturer Warranty', 'Duration (Year)']] = osg_df['Retailer SKU'].apply(
            lambda sku: pd.Series(extract_warranty_duration(sku))
        )
        def highlight_row(row):
            missing_fields = pd.isna(row.get('Model')) or str(row.get('Model')).strip() == ''
            missing_fields |= pd.isna(row.get('IMEI')) or str(row.get('IMEI')).strip() == ''
            try:
                if float(row.get('Plan Price', 0)) < 0:
                    missing_fields |= True
            except:
                missing_fields |= True
            return ['background-color: lightblue'] * len(row) if missing_fields else [''] * len(row)
        

        final_columns = [
            'Customer Mobile', 'Date', 'Invoice Number','Product Invoice Number', 'Customer Name', 'Store Code', 'Branch', 'Region',
            'IMEI', 'Category', 'Brand', 'Quantity', 'Item Code', 'Model', 'Plan Type', 'EWS QTY', 'Item Rate',
            'Plan Price', 'Sold Price', 'Email', 'Product Count', 'Manufacturer Warranty', 'Retailer SKU', 'OnsiteGo SKU',
            'Duration (Year)', 'Total Coverage', 'Comment', 'Return Flag', 'Return against invoice No.',
            'Primary Invoice No.'
        ]

        for col in final_columns:
            if col not in osg_df.columns:
                osg_df[col] = ''
        osg_df['Quantity'] = 1
        osg_df['EWS QTY'] = 1
        osg_df = osg_df[final_columns]
        

        st.success("✅ Data Mapping Completed")
        

        @st.cache_data
        def convert_df(df):
           output = io.BytesIO()
           styled_df = df.style.apply(highlight_row, axis=1)
           with pd.ExcelWriter(output, engine='openpyxl') as writer:
            styled_df.to_excel(writer, index=False)
           output.seek(0)
           return output
        
        
        excel_data = convert_df(osg_df)

        st.download_button(
          label="Download Excel Report",
          data=excel_data,
          file_name="report.xlsx",
          mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

           
        
    else:
        st.info("Please upload both required files.")
